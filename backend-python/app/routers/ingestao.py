import io
import logging
import random
from datetime import date, datetime, timedelta

import openpyxl
from fastapi import APIRouter, UploadFile, File, HTTPException

from app.core.supabase_client import get_supabase

logger = logging.getLogger(__name__)
router = APIRouter()

# Max upload size: 10MB
MAX_UPLOAD_BYTES = 10 * 1024 * 1024


@router.post("/excel")
async def importar_excel(file: UploadFile = File(...)):
    """
    Importa dados de uma planilha Excel (.xlsx)
    Colunas esperadas: cliente, horario_planejado, peso_previsto_kg, turno
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Arquivo deve ser .xlsx ou .xls")

    try:
        content = await file.read()

        # Fix #18: File size limit
        if len(content) > MAX_UPLOAD_BYTES:
            raise HTTPException(400, f"Arquivo excede o limite de {MAX_UPLOAD_BYTES // (1024*1024)}MB")

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            raise HTTPException(400, "Planilha vazia ou sem linhas de dados")

        raw_headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        header_map = {h: idx for idx, h in enumerate(raw_headers) if h}

        required = {"cliente", "horario_planejado"}
        if not required.issubset(set(header_map.keys())):
            raise HTTPException(
                400,
                f"Colunas obrigatórias: {required}. Encontradas: {list(header_map.keys())}",
            )

        sb = get_supabase()
        today = date.today().isoformat()

        payloads = []
        for row in rows[1:]:
            if not row or not any(v is not None and str(v).strip() for v in row):
                continue
            
            cliente_val = row[header_map["cliente"]] if "cliente" in header_map and header_map["cliente"] < len(row) else None
            horario_val = row[header_map["horario_planejado"]] if "horario_planejado" in header_map and header_map["horario_planejado"] < len(row) else None
            
            if cliente_val is None or horario_val is None:
                continue

            peso_val = 0.0
            if "peso_previsto_kg" in header_map and header_map["peso_previsto_kg"] < len(row):
                try:
                    peso_val = float(row[header_map["peso_previsto_kg"]] or 0)
                except (ValueError, TypeError):
                    peso_val = 0.0

            turno_val = "manha"
            if "turno" in header_map and header_map["turno"] < len(row) and row[header_map["turno"]]:
                turno_val = str(row[header_map["turno"]]).strip()

            # Formata horario se for datetime/time
            horario_str = str(horario_val).strip()
            if isinstance(horario_val, datetime):
                horario_str = horario_val.strftime("%H:%M")

            payloads.append(
                {
                    "cliente": str(cliente_val).strip(),
                    "horario_planejado": horario_str,
                    "peso_previsto_kg": peso_val,
                    "turno": turno_val,
                    "data": today,
                    "status": "pendente",
                    "planta": "CNN",
                }
            )

        if payloads:
            sb.table("expedicao_diaria").insert(payloads).execute()

        return {
            "success": True,
            "message": f"✅ {len(payloads)} expedições importadas do Excel",
            "data": {"total_importados": len(payloads), "arquivo": file.filename},
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Erro ao processar Excel: %s", e)
        raise HTTPException(500, f"Erro ao processar Excel: {str(e)}")


@router.post("/excel-producao")
async def importar_excel_producao(file: UploadFile = File(...)):
    """
    Importa dados de produção de uma planilha Excel
    Colunas: cliente, tipo (entrada/saida), peso_kg, turno
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Arquivo deve ser .xlsx ou .xls")

    try:
        content = await file.read()

        # Fix #18: File size limit
        if len(content) > MAX_UPLOAD_BYTES:
            raise HTTPException(400, f"Arquivo excede o limite de {MAX_UPLOAD_BYTES // (1024*1024)}MB")

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            raise HTTPException(400, "Planilha vazia ou sem linhas de dados")

        raw_headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        header_map = {h: idx for idx, h in enumerate(raw_headers) if h}

        required = {"tipo", "peso_kg"}
        if not required.issubset(set(header_map.keys())):
            raise HTTPException(400, f"Colunas obrigatórias: {required}")

        sb = get_supabase()
        today = date.today().isoformat()

        payloads = []
        for row in rows[1:]:
            if not row or not any(v is not None and str(v).strip() for v in row):
                continue

            tipo_val = row[header_map["tipo"]] if "tipo" in header_map and header_map["tipo"] < len(row) else None
            peso_val = row[header_map["peso_kg"]] if "peso_kg" in header_map and header_map["peso_kg"] < len(row) else None

            if tipo_val is None or peso_val is None:
                continue

            try:
                peso_float = float(peso_val)
            except (ValueError, TypeError):
                continue

            cliente_val = "Geral Planta"
            if "cliente" in header_map and header_map["cliente"] < len(row) and row[header_map["cliente"]]:
                cliente_val = str(row[header_map["cliente"]]).strip()

            turno_val = "manha"
            if "turno" in header_map and header_map["turno"] < len(row) and row[header_map["turno"]]:
                turno_val = str(row[header_map["turno"]]).strip()

            payloads.append(
                {
                    "cliente": cliente_val,
                    "tipo": str(tipo_val).strip(),
                    "peso_kg": peso_float,
                    "turno": turno_val,
                    "data": today,
                    "origem": "excel",
                    "planta": "CNN",
                }
            )

        if payloads:
            sb.table("producao_diaria").insert(payloads).execute()

        return {
            "success": True,
            "message": f"✅ {len(payloads)} registros de produção importados",
            "data": {"total_importados": len(payloads)},
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Erro ao processar Excel de produção: %s", e)
        raise HTTPException(500, f"Erro ao processar Excel: {str(e)}")


@router.post("/mock-coala")
async def mock_coala_data():
    """
    Simula um servidor COALA gerando dados mock para o dia atual
    com >10 toneladas e perfis de hospital/indústria.
    """
    sb = get_supabase()
    today = date.today().isoformat()
    now = datetime.now()

    # Check if already has data today
    check = sb.table("expedicao_diaria").select("id").eq("data", today).limit(1).execute()
    if check.data:
        return {
            "success": True,
            "message": "⚠️ Já existem dados para hoje. Nenhum dado mock inserido.",
        }

    CLIENTS = [
        "Hospital São Lucas",
        "Hospital das Clínicas",
        "Rede de Hotéis Plaza",
        "Indústria Automotiva XYZ",
        "Hospital Santa Maria",
        "Resort Internacional Sul",
        "Indústria Farmacêutica Vida",
        "Maternidade Central",
        "Hotel Executivo Faria Lima",
        "Metalúrgica Nacional",
        "Hospital de Câncer Esperança",
        "Rede D'Or São Luiz",
        "Indústria Química Apex",
        "Hotel Ibis Paulista",
        "Hospital Sírio Libanês",
    ]

    expedicoes = []
    producoes = []

    for _ in range(15):
        cliente = random.choice(CLIENTS)
        CLIENTS.remove(cliente)

        peso_prev = random.randint(1500, 4000)
        status_choice = random.choice(["concluido", "no_prazo", "proximo", "atrasado", "pendente"])

        if status_choice == "concluido":
            h_plan = (now - timedelta(minutes=random.randint(60, 300))).strftime("%H:%M")
            h_real = (now - timedelta(minutes=random.randint(65, 310))).strftime("%H:%M")
            peso_exp = peso_prev
        elif status_choice == "atrasado":
            h_plan = (now - timedelta(minutes=random.randint(30, 120))).strftime("%H:%M")
            h_real = None
            peso_exp = 0
        elif status_choice == "proximo":
            h_plan = (now + timedelta(minutes=random.randint(5, 25))).strftime("%H:%M")
            h_real = None
            peso_exp = 0
        else:
            h_plan = (now + timedelta(minutes=random.randint(40, 200))).strftime("%H:%M")
            h_real = None
            peso_exp = 0

        expedicoes.append(
            {
                "cliente": cliente,
                "horario_planejado": h_plan,
                "horario_real": h_real,
                "status": status_choice,
                "data": today,
                "peso_previsto_kg": peso_prev,
                "peso_expedido_kg": peso_exp,
                "turno": "manha" if random.random() > 0.3 else "tarde",
                "planta": "CNN",
            }
        )

    # Batch insert expedições
    if expedicoes:
        sb.table("expedicao_diaria").insert(expedicoes).execute()

    # Produção
    entrada_base = random.randint(22000, 28000)
    producoes.append(
        {
            "cliente": "Geral Planta",
            "tipo": "entrada",
            "peso_kg": entrada_base,
            "turno": "manha",
            "data": today,
            "origem": "api",
            "planta": "CNN",
        }
    )

    for exp in expedicoes:
        if exp["status"] == "concluido":
            producoes.append(
                {
                    "cliente": exp["cliente"],
                    "tipo": "saida",
                    "peso_kg": exp["peso_expedido_kg"],
                    "turno": exp["turno"],
                    "data": today,
                    "origem": "api",
                    "planta": "CNN",
                }
            )

    if producoes:
        sb.table("producao_diaria").insert(producoes).execute()

    return {
        "success": True,
        "message": f"✅ Mock COALA: 15 expedições ({entrada_base}kg processados) inseridos",
    }


@router.post("/mock-historico")
async def mock_historico_data():
    """
    Simula dados dos últimos 30 dias para preencher o Histórico e Comparativos.
    """
    sb = get_supabase()
    today = datetime.now()

    # Fix #12: Single query to find existing dates instead of 30 individual queries
    thirty_ago = (today - timedelta(days=30)).date().isoformat()
    existing_prod = (
        sb.table("producao_diaria")
        .select("data")
        .gte("data", thirty_ago)
        .execute()
    )
    existing_dates = set(r["data"] for r in (existing_prod.data or []))

    expedicoes = []
    producoes = []

    CLIENTS = [
        "Hospital São Lucas",
        "Hospital das Clínicas",
        "Rede de Hotéis Plaza",
        "Indústria Automotiva XYZ",
        "Hospital Santa Maria",
    ]

    # Fix #3: 30 days, not 300
    for i in range(1, 31):
        dt = (today - timedelta(days=i)).date().isoformat()

        # Skip dates that already have data
        if dt in existing_dates:
            continue

        # Producao
        entrada = random.randint(22000, 28000)
        eficiencia = random.uniform(0.85, 0.99)
        saida = int(entrada * eficiencia)

        producoes.append(
            {
                "cliente": "Geral Planta",
                "tipo": "entrada",
                "peso_kg": entrada,
                "turno": "manha",
                "data": dt,
                "origem": "api",
                "planta": "CNN",
            }
        )
        producoes.append(
            {
                "cliente": "Geral Planta",
                "tipo": "saida",
                "peso_kg": saida,
                "turno": "manha",
                "data": dt,
                "origem": "api",
                "planta": "CNN",
            }
        )

        # Expedicao
        for _ in range(random.randint(5, 12)):
            status_choice = "concluido" if random.random() > 0.15 else "atrasado"
            expedicoes.append(
                {
                    "cliente": random.choice(CLIENTS),
                    "horario_planejado": "12:00",
                    "horario_real": "11:50" if status_choice == "concluido" else None,
                    "status": status_choice,
                    "data": dt,
                    "peso_previsto_kg": random.randint(1500, 4000),
                    "peso_expedido_kg": random.randint(1500, 4000) if status_choice == "concluido" else 0,
                    "turno": "manha",
                    "planta": "CNN",
                }
            )

    # Batch inserts
    if producoes:
        sb.table("producao_diaria").insert(producoes).execute()
    if expedicoes:
        sb.table("expedicao_diaria").insert(expedicoes).execute()

    return {
        "success": True,
        "message": "✅ Dados de Histórico mockados para os últimos 30 dias!",
    }
